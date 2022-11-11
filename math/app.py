import openpyxl

path = "input.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj =wb_obj.active

start_x = 10
start_y = 10

start_Cell_val = [6, "a" , 1, 1]

sheet_obj.cell(row= start_y , column= start_x).value = str(start_Cell_val[0]) + "(" +start_Cell_val[1]+ str(start_Cell_val[2]) + ")" + "+" + str(start_Cell_val[3])

curent_x = 10
curent_y = 10


for i in range(1,5):
    
    current_sell = sheet_obj.cell(row = curent_y  , column = curent_x )
    c= [int(current_sell.value[0]) , "a" , int(current_sell.value[current_sell.value.index("a") + 1 : current_sell.value.index(")")]) , int(current_sell.value[current_sell.value.index(")") +1 : len(current_sell.value) ])]    
    sheet_obj.cell(row=curent_x,column=curent_y+i).value = str(c[0]) + "(" + c[1]+ str( c[2]) + ")" + "+" + str( c[3])
    


   
    



wb_obj.save(filename="input-inp.xlsx")